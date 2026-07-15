"""
Excel to Images Converter
Converts Excel sheets to PNG images to reduce token usage when reading large files.
"""

import argparse
import json
import os
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from PIL import Image, ImageDraw, ImageFont
except ImportError as e:
    print(f"Error: Required package not found: {e}")
    print("Please install required packages: pip install openpyxl pillow")
    sys.exit(1)


class ExcelToImageConverter:
    """Convert Excel sheets to images"""
    
    def __init__(self, excel_file, output_dir=None, dpi=250, max_width=1920, font_size=14, show_gridlines=False):
        self.excel_file = Path(excel_file)
        if not self.excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_file}")
        
        # Output directory: same location as source file with _images suffix
        if output_dir is None:
            self.output_dir = self.excel_file.parent / f"{self.excel_file.stem}_images"
        else:
            self.output_dir = Path(output_dir)
        
        self.dpi = dpi
        self.max_width = max_width
        self.cell_width = 100
        self.cell_height = 25
        self.font_size = font_size
        self.show_gridlines = show_gridlines
        
        # Create output directory
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def get_cell_dimensions(self, ws):
        """Calculate actual cell dimensions based on column widths and row heights"""
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Get column widths
        col_widths = {}
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            width = ws.column_dimensions[col_letter].width
            if width is None:
                width = 8.43  # Default Excel column width
            col_widths[col] = int(width * 12)  # Maximum spacing for readability
        
        # Get row heights
        row_heights = {}
        for row in range(1, max_row + 1):
            height = ws.row_dimensions[row].height
            if height is None:
                height = 15  # Default Excel row height
            # Very large multiplier to completely prevent text overlap
            row_heights[row] = max(int(height * 3.5), self.font_size + 24)
        
        return col_widths, row_heights
    
    def render_sheet_to_image(self, ws, sheet_name):
        """Render a worksheet to an image"""
        print(f"  Processing sheet: {sheet_name}")
        
        max_row = ws.max_row
        max_col = ws.max_column
        
        if max_row == 0 or max_col == 0:
            print(f"  Warning: Sheet '{sheet_name}' is empty, skipping")
            return None
        
        # Get cell dimensions
        col_widths, row_heights = self.get_cell_dimensions(ws)
        
        # Calculate image size
        img_width = sum(col_widths.values()) + 10
        img_height = sum(row_heights.values()) + 10
        
        # Limit image width
        if img_width > self.max_width:
            scale_factor = self.max_width / img_width
            img_width = self.max_width
            img_height = int(img_height * scale_factor)
            for col in col_widths:
                col_widths[col] = int(col_widths[col] * scale_factor)
            for row in row_heights:
                row_heights[row] = int(row_heights[row] * scale_factor)
        
        # Create image
        img = Image.new('RGB', (img_width, img_height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Try to load a font (with Japanese support)
        try:
            # Windows Japanese fonts
            font = ImageFont.truetype("C:/Windows/Fonts/msgothic.ttc", self.font_size)
        except:
            try:
                font = ImageFont.truetype("C:/Windows/Fonts/meiryo.ttc", self.font_size)
            except:
                try:
                    font = ImageFont.truetype("arial.ttf", self.font_size)
                except:
                    try:
                        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", self.font_size)
                    except:
                        font = ImageFont.load_default()
        
        # Draw cells
        y_offset = 5
        for row in range(1, max_row + 1):
            x_offset = 5
            row_height = row_heights[row]
            
            for col in range(1, max_col + 1):
                col_width = col_widths[col]
                cell = ws.cell(row=row, column=col)
                
                # Draw cell border (optional)
                if self.show_gridlines:
                    draw.rectangle(
                        [x_offset, y_offset, x_offset + col_width, y_offset + row_height],
                        outline='lightgray',
                        width=1
                    )
                
                # Draw cell value
                if cell.value is not None:
                    text = str(cell.value)
                    # Truncate long text
                    if len(text) > 50:
                        text = text[:47] + "..."
                    
                    # Draw text with better padding and vertical centering
                    text_x = x_offset + 5
                    # Center text vertically in cell
                    text_y = y_offset + (row_height - self.font_size) // 2
                    draw.text((text_x, text_y), text, fill='black', font=font)
                
                x_offset += col_width
            
            y_offset += row_height
        
        # Save image
        output_file = self.output_dir / f"{sheet_name}.png"
        img.save(output_file, 'PNG')
        print(f"  Saved: {output_file}")
        
        return output_file
    
    def convert(self, sheet_names=None):
        """Convert Excel file to images"""
        print(f"Loading Excel file: {self.excel_file}")
        
        try:
            wb = load_workbook(self.excel_file, data_only=True)
        except Exception as e:
            print(f"Error loading Excel file: {e}")
            return {"error": str(e)}
        
        # Get sheets to process
        if sheet_names:
            sheets_to_process = [s for s in sheet_names if s in wb.sheetnames]
            if not sheets_to_process:
                print(f"Warning: None of the specified sheets found in workbook")
                sheets_to_process = wb.sheetnames
        else:
            sheets_to_process = wb.sheetnames
        
        print(f"Found {len(sheets_to_process)} sheet(s) to process")
        
        # Convert each sheet
        results = []
        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            output_file = self.render_sheet_to_image(ws, sheet_name)
            if output_file:
                results.append({
                    "sheet": sheet_name,
                    "image": str(output_file),
                    "rows": ws.max_row,
                    "columns": ws.max_column
                })
        
        wb.close()
        
        return {
            "status": "success",
            "excel_file": str(self.excel_file),
            "output_dir": str(self.output_dir),
            "sheets_processed": len(results),
            "images": results
        }


def main():
    parser = argparse.ArgumentParser(
        description="Convert Excel sheets to PNG images to reduce token usage"
    )
    parser.add_argument(
        "excel_file",
        help="Path to Excel file (.xlsx)"
    )
    parser.add_argument(
        "--output",
        help="Output directory (default: same location as source file with _images suffix)"
    )
    parser.add_argument(
        "--dpi",
        type=int,
        default=150,
        help="Image resolution (default: 150)"
    )
    parser.add_argument(
        "--max-width",
        type=int,
        default=1920,
        help="Maximum image width in pixels (default: 1920)"
    )
    parser.add_argument(
        "--font-size",
        type=int,
        default=12,
        help="Font size for cell text (default: 12)"
    )
    parser.add_argument(
        "--show-gridlines",
        action="store_true",
        help="Show cell gridlines in output images (default: False)"
    )
    parser.add_argument(
        "--sheets",
        nargs="+",
        help="Specific sheet names to convert (default: all sheets)"
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="Output results as JSON"
    )
    
    args = parser.parse_args()
    
    try:
        converter = ExcelToImageConverter(
            args.excel_file,
            output_dir=args.output,
            dpi=args.dpi,
            max_width=args.max_width,
            font_size=args.font_size,
            show_gridlines=args.show_gridlines
        )
        
        result = converter.convert(sheet_names=args.sheets)
        
        if args.json:
            print(json.dumps(result, indent=2))
        else:
            if "error" in result:
                print(f"\nError: {result['error']}")
                sys.exit(1)
            else:
                print(f"\nConversion complete!")
                print(f"   Excel file: {result['excel_file']}")
                print(f"   Output directory: {result['output_dir']}")
                print(f"   Sheets processed: {result['sheets_processed']}")
                print(f"\nGenerated images:")
                for img in result['images']:
                    print(f"   - {img['sheet']}: {img['rows']}x{img['columns']} cells -> {img['image']}")
    
    except Exception as e:
        error_msg = str(e).encode('utf-8', errors='replace').decode('utf-8')
        print(f"Error: {error_msg}")
        sys.exit(1)


if __name__ == "__main__":
    main()
